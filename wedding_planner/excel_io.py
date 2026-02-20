import openpyxl
from openpyxl import Workbook
from .models import SeatingPlan, Guest, Table

class ExcelIO:
    @staticmethod
    def save_to_xlsx(seating_plan: SeatingPlan, filename: str):
        wb = Workbook()
        
        # Sheet 1: Guests (Headers first, we will populate guests later)
        ws_guests = wb.active
        ws_guests.title = "Guests"
        ws_guests.append(["ID", "Name", "Category", "Capacity", "Table ID"])
        
        # Sheet 2: Tables
        ws_tables = wb.create_sheet("Tables")
        ws_tables.append(["ID", "Name", "Capacity", "X", "Y"])
        
        table_row_map = {}
        for idx, table in enumerate(seating_plan.tables.values(), start=2):
            ws_tables.append([table.id, table.name, table.capacity, table.x, table.y])
            table_row_map[table.id] = idx
            
        # Write Guests
        for guest in seating_plan.guests.values():
            if guest.table_id and guest.table_id in table_row_map:
                row_idx = table_row_map[guest.table_id]
                table_val = f"=Tables!A{row_idx}"
            else:
                table_val = guest.table_id if guest.table_id else ""
            ws_guests.append([guest.id, guest.name, guest.category, guest.size, table_val])

        # Sheet 3: Metadata (IDs)
        ws_meta = wb.create_sheet("Metadata")
        ws_meta.append(["Next Guest ID", "Next Table ID"])
        ws_meta.append([seating_plan.next_guest_id, seating_plan.next_table_id])

        wb.save(filename)

    @staticmethod
    def load_from_xlsx(filename: str, seating_plan: SeatingPlan, clear: bool = True):
        wb = openpyxl.load_workbook(filename, data_only=True)
        
        if clear:
            # Clear existing data
            seating_plan.guests.clear()
            seating_plan.tables.clear()
            
        table_mapping = {} # old_id -> new_id
        guest_mapping = {} # old_id -> new_id
        
        # Load Tables
        if "Tables" in wb.sheetnames:
            ws_tables = wb["Tables"]
            for row in ws_tables.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    t_id_raw, name, capacity, x, y = row[:5]
                    
                    try:
                        t_id = int(float(t_id_raw))
                    except (ValueError, TypeError):
                        continue
                    # Ensure numeric values are ints
                    try:
                        capacity = int(float(capacity))
                        x = int(float(x))
                        y = int(float(y))
                    except (ValueError, TypeError):
                        pass # Keep original or handle error? For now, let it fail or assume valid input if not caught here, but int() conversion is what we want.
                    
                    # Merge handling
                    old_t_id = t_id
                    if not clear and t_id in seating_plan.tables:
                        t_id = seating_plan.next_table_id
                        seating_plan.next_table_id += 1
                        
                    table_mapping[old_t_id] = t_id
                    
                    table = Table(id=t_id, name=name, capacity=capacity, x=x, y=y)
                    seating_plan.tables[t_id] = table

        # Load Guests
        if "Guests" in wb.sheetnames:
            ws_guests = wb["Guests"]
            for row in ws_guests.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    # Check row length to support backward compatibility (old files have 4 cols, new have 5)
                    if len(row) >= 5:
                        g_id, name, category, size, table_id = row[:5]
                        if size is None: size = 1
                    else:
                        g_id, name, category, table_id = row[:4]
                        size = 1
                    
                    try:
                        g_id = int(float(g_id))
                    except (ValueError, TypeError):
                        continue
                    
                    if table_id is not None and str(table_id).strip() != "":
                         try:
                              table_id = int(float(table_id))
                              if not clear and table_id in table_mapping:
                                  table_id = table_mapping[table_id]
                         except (ValueError, TypeError):
                              table_id = None
                    else:
                         table_id = None
                    
                    old_g_id = g_id
                    if not clear and g_id in seating_plan.guests:
                        g_id = seating_plan.next_guest_id
                        seating_plan.next_guest_id += 1
                        
                    guest_mapping[old_g_id] = g_id
                    
                    guest = Guest(id=g_id, name=name, category=category, size=int(size), table_id=table_id)
                    seating_plan.guests[g_id] = guest
                    
                    # Link to table
                    if table_id is not None and table_id in seating_plan.tables:
                        seating_plan.tables[table_id].guest_ids.append(g_id)

        # Load Metadata
        if "Metadata" in wb.sheetnames:
            ws_meta = wb["Metadata"]
            row_meta = next(ws_meta.iter_rows(min_row=2, values_only=True), None)
            if row_meta:
                loaded_next_guest_id = row_meta[0]
                loaded_next_table_id = row_meta[1]
                if clear:
                    seating_plan.next_guest_id = loaded_next_guest_id
                    seating_plan.next_table_id = loaded_next_table_id
                else:
                    try:
                        seating_plan.next_guest_id = max(seating_plan.next_guest_id, int(loaded_next_guest_id))
                        seating_plan.next_table_id = max(seating_plan.next_table_id, int(loaded_next_table_id))
                    except:
                        pass
        
        # Recalculate IDs regardless to be safe
        if seating_plan.guests:
            seating_plan.next_guest_id = max(seating_plan.next_guest_id, max(seating_plan.guests.keys()) + 1)
        if seating_plan.tables:
            seating_plan.next_table_id = max(seating_plan.next_table_id, max(seating_plan.tables.keys()) + 1)

    @staticmethod
    def get_headers(filename: str) -> list[str]:
        """Returns the headers (first row) of the active sheet."""
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell) for cell in row if cell is not None]
            break
        wb.close()
        return headers

    @staticmethod
    def import_groups_to_plan(filename: str, group_col: str, count_col: str, seating_plan: SeatingPlan, category_col: str = None):
        """
        Imports groups from Excel.
        group_col: Header name for the group/guest name column
        count_col: Header name for the count column
        category_col: Optional header name for the category column
        """
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        ws = wb.active
        
        # Find column indices
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell) for cell in row]
            break
            
        try:
            group_idx = headers.index(group_col)
            count_idx = headers.index(count_col)
            category_idx = headers.index(category_col) if category_col else None
        except ValueError as e:
             raise ValueError(f"Column not found in headers: {headers}. Error: {e}")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            
            group_name = row[group_idx]
            if not group_name: continue
            
            try:
                count = int(row[count_idx])
            except (ValueError, TypeError):
                # meaningful default or skip? Let's assume 1 if invalid
                count = 1
            
            if count <= 0: continue
            
            # Determine category
            category = str(group_name)
            if category_idx is not None:
                cat_val = row[category_idx]
                if cat_val:
                    category = str(cat_val)

            # Create guests
            seating_plan.add_guest(name=str(group_name), category=category, size=count)
        
        wb.close()
