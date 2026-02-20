
import sys
import os
import openpyxl
from openpyxl import Workbook

# Add project root
sys.path.append(os.getcwd())

from wedding_planner.models import SeatingPlan
from wedding_planner.excel_io import ExcelIO

def create_buggy_excel(filename):
    wb = Workbook()
    ws = wb.create_sheet("Tables")
    ws.append(["ID", "Name", "Capacity", "X", "Y"])
    # Note: 10.0 is a float
    ws.append([1, "Table 1", 10.0, 100.0, 100.0])
    
    # Also removing the default sheet created properly 
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    wb.save(filename)
    print(f"Created {filename} with float capacity")

def test_load():
    filename = "buggy_tables.xlsx"
    create_buggy_excel(filename)
    
    plan = SeatingPlan()
    try:
        ExcelIO.load_from_xlsx(filename, plan)
        table = plan.tables[1]
        print(f"Table capacity type: {type(table.capacity)}")
        print(f"Table capacity value: {table.capacity}")
        
        # This is where it would crash in GUI
        try:
            print("Attempting range(table.capacity)...")
            _ = list(range(table.capacity))
            print("range() worked!")
        except TypeError as e:
            print(f"Caught expected error: {e}")

    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        if os.path.exists(filename):
            os.remove(filename)

if __name__ == "__main__":
    test_load()
