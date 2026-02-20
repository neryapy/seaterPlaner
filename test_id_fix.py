
import sys
import os
import openpyxl
from openpyxl import Workbook

# Add project root
sys.path.append(os.getcwd())

from wedding_planner.models import SeatingPlan, Table, Guest
from wedding_planner.excel_io import ExcelIO

def create_string_id_excel(filename):
    wb = Workbook()
    ws = wb.create_sheet("Tables")
    ws.append(["ID", "Name", "Capacity", "X", "Y"])
    # "1" as string
    ws.append(["1", "Table 1", 10, 100, 100])
    
    ws_g = wb.create_sheet("Guests")
    ws_g.append(["ID", "Name", "Category", "Size", "Table ID"])
    # "1" as string
    ws_g.append(["1", "Guest 1", "General", 1, "1"])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    wb.save(filename)
    print(f"Created {filename} with string IDs")

def test_load():
    filename = "string_ids.xlsx"
    create_string_id_excel(filename)
    
    plan = SeatingPlan()
    try:
        ExcelIO.load_from_xlsx(filename, plan)
        
        # Check Table ID
        if 1 in plan.tables:
             print("Table ID 1 found (int key). OK.")
        elif "1" in plan.tables:
             print("FAIL: Table ID 1 found as string key.")
        else:
             print("FAIL: Table ID 1 not found.")
             
        table = plan.tables.get(1)
        if table:
            print(f"Table ID type: {type(table.id)}")

        # Check Guest ID and Foreign Key
        if 1 in plan.guests:
             print("Guest ID 1 found (int key). OK.")
        
        guest = plan.guests.get(1)
        if guest:
             print(f"Guest ID type: {type(guest.id)}")
             print(f"Guest Table ID type: {type(guest.table_id)}")
             print(f"Guest Table ID value: {guest.table_id}")
             
             if guest.table_id == 1:
                 print("Guest linked correctly to Table 1 (int). OK.")
             else:
                 print(f"FAIL: Guest linked to {guest.table_id}")

    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(filename):
            os.remove(filename)

if __name__ == "__main__":
    test_load()
